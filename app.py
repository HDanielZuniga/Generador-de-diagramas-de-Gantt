from flask import Flask, render_template_string, request, redirect, url_for, flash, send_file
import plotly.figure_factory as ff
from datetime import datetime
import io
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

app = Flask(__name__)
app.secret_key = "supersecreto"  # Necesario para mensajes flash

# Lista global de tareas con información de planificación y colaboradores
tasks = [
    {
        'Task': 'Análisis',
        'Start': '2025-04-01',
        'Finish': '2025-04-05',
        'Resource': 'Análisis',
        'Collaborators': [
            {"name": "Colaborador 1", "role": "Analista"},
            {"name": "Colaborador 2", "role": "Supervisor"}
        ]
    },
    {
        'Task': 'Diseño',
        'Start': '2025-04-06',
        'Finish': '2025-04-10',
        'Resource': 'Diseño',
        'Collaborators': [
            {"name": "Colaborador 1", "role": "Diseñador"}
        ]
    },
    {
        'Task': 'Implementación (Desarrollo)',
        'Start': '2025-04-11',
        'Finish': '2025-04-20',
        'Resource': 'Desarrollo',
        'Collaborators': [
            {"name": "Colaborador 1", "role": "Desarrollador"},
            {"name": "Colaborador 2", "role": "Tester"}
        ]
    },
    {
        'Task': 'Pruebas',
        'Start': '2025-04-21',
        'Finish': '2025-04-25',
        'Resource': 'Pruebas',
        'Collaborators': [
            {"name": "Colaborador 1", "role": "Tester"},
            {"name": "Colaborador 2", "role": "Verificador"}
        ]
    }
]

def generar_gantt_html(tasks_list, group=True):
    """
    Genera el diagrama de Gantt en HTML usando Plotly.
    (Este gráfico se usa para la vista web, no para el Excel.)
    """
    tareas_ordenadas = sorted(tasks_list, key=lambda t: datetime.strptime(t['Start'], "%Y-%m-%d"))
    fig = ff.create_gantt(tareas_ordenadas, index_col='Resource', show_colorbar=True, group_tasks=group)
    return fig.to_html(full_html=False)

def validar_fecha(fecha_texto):
    try:
        datetime.strptime(fecha_texto, "%Y-%m-%d")
        return True
    except ValueError:
        return False

@app.route("/")
def index():
    if tasks:
        gantt_chart = generar_gantt_html(tasks, group=True)
    else:
        gantt_chart = "<p>No hay tareas para mostrar. Por favor, agrega una tarea.</p>"
    
    html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Planificación y Diagrama de Gantt</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; }
            .container { display: flex; flex-wrap: wrap; }
            .chart, .table { flex: 1; min-width: 300px; padding: 10px; }
            .chart { border-right: 1px solid #ccc; }
            table { border-collapse: collapse; width: 100%; margin-top: 10px; }
            th, td { border: 1px solid #999; padding: 8px; text-align: center; }
            th { background-color: #eee; }
            ul { list-style-type: none; padding: 0; margin: 0; }
            .actions { margin-bottom: 20px; }
            .actions a { margin-right: 20px; }
        </style>
    </head>
    <body>
        <h1>Planificación del Proyecto</h1>
        <div class="actions">
            <a href="{{ url_for('add_task') }}">Agregar Tarea</a>
            <a href="{{ url_for('delete_all') }}">Eliminar Todas las Tareas</a>
            <a href="{{ url_for('download_excel') }}">Descargar Excel</a>
        </div>
        {% with messages = get_flashed_messages() %}
          {% if messages %}
            <ul style="color:red;">
            {% for message in messages %}
              <li>{{ message }}</li>
            {% endfor %}
            </ul>
          {% endif %}
        {% endwith %}
        <div class="container">
            <div class="chart">
                <h2>Diagrama de Gantt (Vista Web)</h2>
                {{ gantt_chart|safe }}
            </div>
            <div class="table">
                <h2>Resumen de la Planificación</h2>
                <table>
                    <tr>
                        <th>Tarea</th>
                        <th>Inicio</th>
                        <th>Fin</th>
                        <th>Fase/Grupo</th>
                        <th>Colaboradores y Roles</th>
                    </tr>
                    {% for task in tasks %}
                    <tr>
                        <td>{{ task['Task'] }}</td>
                        <td>{{ task['Start'] }}</td>
                        <td>{{ task['Finish'] }}</td>
                        <td>{{ task['Resource'] }}</td>
                        <td>
                            <ul>
                            {% for collab in task['Collaborators'] %}
                                <li>{{ collab.name }} (<em>{{ collab.role }}</em>)</li>
                            {% endfor %}
                            </ul>
                        </td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
        </div>
    </body>
    </html>
    """
    return render_template_string(html_template, gantt_chart=gantt_chart, tasks=tasks)

@app.route("/add", methods=["GET", "POST"])
def add_task():
    if request.method == "POST":
        task = request.form.get("Task")
        start = request.form.get("Start")
        finish = request.form.get("Finish")
        resource = request.form.get("Resource")
        
        if not (validar_fecha(start) and validar_fecha(finish)):
            flash("Las fechas deben tener el formato YYYY-MM-DD.")
            return redirect(url_for("add_task"))
        if datetime.strptime(finish, "%Y-%m-%d") < datetime.strptime(start, "%Y-%m-%d"):
            flash("La fecha de fin debe ser posterior a la fecha de inicio.")
            return redirect(url_for("add_task"))
        
        names = request.form.getlist("collaborator_name[]")
        roles = request.form.getlist("collaborator_role[]")
        if not names or not roles or len(names) != len(roles):
            flash("Debe proporcionar la información de todos los colaboradores.")
            return redirect(url_for("add_task"))
        
        collaborators = [{"name": names[i], "role": roles[i]} for i in range(len(names))]
        
        new_task = {
            'Task': task,
            'Start': start,
            'Finish': finish,
            'Resource': resource,
            'Collaborators': collaborators
        }
        tasks.append(new_task)
        return redirect(url_for("index"))
    
    form_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Agregar Tarea al Proyecto</title>
        <meta charset="utf-8">
        <style>
            .collaborator-box { margin-bottom: 10px; border: 1px solid #ccc; padding: 10px; }
        </style>
    </head>
    <body>
        <h1>Agregar Tarea al Proyecto</h1>
        <form method="POST" action="{{ url_for('add_task') }}">
            <label for="Task">Tarea:</label>
            <input type="text" id="Task" name="Task" required><br><br>
            
            <label for="Start">Inicio (YYYY-MM-DD):</label>
            <input type="text" id="Start" name="Start" required><br><br>
            
            <label for="Finish">Fin (YYYY-MM-DD):</label>
            <input type="text" id="Finish" name="Finish" required><br><br>
            
            <label for="Resource">Fase/Grupo:</label>
            <select id="Resource" name="Resource" required>
                <option value="">Seleccione una fase</option>
                <option value="Análisis">Análisis</option>
                <option value="Diseño">Diseño</option>
                <option value="Desarrollo">Desarrollo</option>
                <option value="Pruebas">Pruebas</option>
            </select><br><br>
            
            <label for="numCollaborators">Número de Colaboradores:</label>
            <input type="number" id="numCollaborators" name="numCollaborators" min="1" required><br><br>
            
            <div id="collaboratorsContainer"></div>
            
            <button type="submit">Agregar</button>
        </form>
        <br>
        <a href="{{ url_for('index') }}">Volver al Diagrama</a>
        
        <script>
        var phaseRoles = {
            "Análisis": ["Analista", "Supervisor"],
            "Diseño": ["Diseñador", "Especialista UX/UI"],
            "Desarrollo": ["Desarrollador", "Tester"],
            "Pruebas": ["Tester", "Verificador"]
        };

        function generateCollaboratorBoxes(){
            var num = parseInt(document.getElementById("numCollaborators").value) || 0;
            var phase = document.getElementById("Resource").value;
            var container = document.getElementById("collaboratorsContainer");
            container.innerHTML = "";
            for (var i = 0; i < num; i++) {
                var div = document.createElement("div");
                div.className = "collaborator-box";
                var html = "<strong>Colaborador " + (i+1) + ":</strong><br>" +
                           "Nombre: <input type='text' name='collaborator_name[]' required>";
                if(phase && phaseRoles.hasOwnProperty(phase)){
                    html += " &nbsp; Rol: <select name='collaborator_role[]' required>";
                    var roles = phaseRoles[phase];
                    for(var j=0; j<roles.length; j++){
                        html += "<option value='"+roles[j]+"'>"+roles[j]+"</option>";
                    }
                    html += "</select>";
                } else {
                    html += " &nbsp; Rol: <input type='text' name='collaborator_role[]' required>";
                }
                div.innerHTML = html;
                container.appendChild(div);
            }
        }

        document.getElementById("numCollaborators").addEventListener("change", generateCollaboratorBoxes);
        document.getElementById("Resource").addEventListener("change", generateCollaboratorBoxes);
        </script>
    </body>
    </html>
    """
    return render_template_string(form_template)

@app.route("/delete_all")
def delete_all():
    tasks.clear()
    flash("Se han eliminado todas las tareas.")
    return redirect(url_for("index"))

@app.route("/download")
def download_excel():
    # Hoja 1: Planificación (tabla de datos)
    data = []
    for task in tasks:
        collaborators_str = ", ".join([f"{collab['name']} ({collab['role']})" for collab in task['Collaborators']])
        data.append({
            "Tarea": task['Task'],
            "Inicio": task['Start'],
            "Fin": task['Finish'],
            "Fase/Grupo": task['Resource'],
            "Colaboradores y Roles": collaborators_str
        })
    df = pd.DataFrame(data)
    
    wb = Workbook()
    
    # Escribir hoja "Planificación"
    ws_data = wb.active
    ws_data.title = "Planificación"
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_data.cell(row=r_idx, column=c_idx, value=value)
    
    # Hoja 2: Generar datos para el gráfico de Gantt
    ws_gantt = wb.create_sheet(title="GanttChart")
    ws_gantt.append(["Tarea", "Offset", "Duración"])
    
    # Calcular la fecha base (mínima) entre los inicios de las tareas
    fechas_inicio = [datetime.strptime(task["Start"], "%Y-%m-%d") for task in tasks]
    base_date = min(fechas_inicio) if fechas_inicio else datetime.today()
    
    # Determinar la duración y el offset para cada tarea
    for task in tasks:
        start_dt = datetime.strptime(task["Start"], "%Y-%m-%d")
        finish_dt = datetime.strptime(task["Finish"], "%Y-%m-%d")
        offset = (start_dt - base_date).days
        # Sumamos 1 para incluir el día final
        duration = (finish_dt - start_dt).days + 1
        ws_gantt.append([task["Task"], offset, duration])
    
    # Crear un gráfico de barras apiladas para simular el diagrama de Gantt
    chart = BarChart()
    chart.type = "bar"
    chart.style = 12
    chart.title = "Diagrama de Gantt"
    chart.y_axis.title = "Tareas"
    chart.x_axis.title = "Días"
    chart.y_axis.reverseOrder = True
    chart.grouping = "stacked"
    
    # Ajustar dimensiones del gráfico
    chart.width = 25  # Ancho del gráfico
    chart.height = 12 # Alto del gráfico
    
    # Ajustar la escala del eje X
    # Determinar el offset máximo para fijar la escala
    max_offset = max((datetime.strptime(t["Finish"], "%Y-%m-%d") - base_date).days + 1 for t in tasks) if tasks else 0
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = max_offset
    chart.x_axis.majorUnit = 1
    chart.x_axis.minorUnit = 1
    
    # Definir referencias para datos y categorías
    data_ref = Reference(ws_gantt, min_col=2, min_row=1, max_col=3, max_row=ws_gantt.max_row)
    categories = Reference(ws_gantt, min_col=1, min_row=2, max_row=ws_gantt.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories)
    
    # Ocultar la primera serie (Offset) para que no se vea
    # ni en el gráfico ni en la leyenda
    if chart.series:
        # Eliminar la leyenda por completo
        chart.legend = None
        
        # Cambiar el color de la segunda serie (Duración) a verde
        if len(chart.series) > 1:
            chart.series[1].graphicalProperties.solidFill = "00B050"
    
    ws_gantt.add_chart(chart, "E5")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name="planificacion_con_grafico.xlsx", as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(debug=True)
